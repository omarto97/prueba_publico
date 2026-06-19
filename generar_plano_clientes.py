#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generador de archivo plano de CLIENTES (tipo de registro 201) para SIESA.

Uso:
    python generar_plano_clientes.py layout.xlsx datos.xlsx salida.txt

- layout.xlsx : el archivo "archivo_plano.xlsx" con las columnas
                NOMBRE | DESCRIPCION | TIPO | OBSERVACIONES | OBL | INICIO | TAMAÑO | TOTAL
- datos.xlsx  : plantilla con una columna por campo (incluye F_NUMERO_REG, F_TIPO_REG, etc.)
                Una fila = un registro de cliente.
- salida.txt  : archivo plano de ancho fijo resultante, codificado en ANSI/cp1252
                (cambia ENCODING_SALIDA si tu conector espera otra cosa).
"""

import sys
import re
import unicodedata
from datetime import datetime
import openpyxl

ENCODING_SALIDA = "cp1252"   # SIESA normalmente espera ANSI/Latin-1. Cambia a "utf-8" si tu conector lo requiere.

# ---------------------------------------------------------------------------
# 1. LECTURA DEL LAYOUT
# ---------------------------------------------------------------------------

def normalizar(nombre):
    """Normaliza nombres de campo: sin tildes, mayúsculas, sin espacios extra."""
    if nombre is None:
        return ""
    s = str(nombre).strip().upper()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s


def leer_layout(path_layout):
    wb = openpyxl.load_workbook(path_layout, data_only=True)
    ws = wb.active
    campos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre, desc, tipo, obs, obl, inicio, tamano, total = (list(row) + [None] * 8)[:8]
        if not nombre or inicio is None or tamano is None:
            continue  # saltar encabezados/secciones/notas
        if not isinstance(inicio, (int, float)) or not isinstance(tamano, (int, float)):
            continue
        campos.append({
            "nombre": normalizar(nombre),
            "nombre_original": str(nombre).strip(),
            "tipo": (tipo or "").strip().lower(),       # 'numérico' / 'alfanumérico'
            "obligatorio": (obl or "").strip().lower(), # 'si' / 'no' / 'dep'
            "obs": obs or "",
            "inicio": int(inicio),
            "tamano": int(tamano),
        })
    campos.sort(key=lambda c: c["inicio"])
    largo_total = max(c["inicio"] + c["tamano"] - 1 for c in campos)
    return campos, largo_total


# ---------------------------------------------------------------------------
# 2. FORMATEO DE VALORES POR CAMPO
# ---------------------------------------------------------------------------

def fmt_decimal_con_signo(valor, ancho):
    """Formato (signo + N enteros + punto + decimales) ej: cupo_credito ancho 21 -> 1+15+1+4."""
    decimales = 4
    enteros = ancho - 1 - 1 - decimales  # signo + punto
    try:
        v = float(valor) if valor not in (None, "") else 0.0
    except ValueError:
        v = 0.0
    signo = "-" if v < 0 else "+"
    v = abs(v)
    s = f"{v:.{decimales}f}"  # ej '123.4500'
    ent, dec = s.split(".")
    ent = ent.zfill(enteros)[-enteros:]
    return f"{signo}{ent}.{dec}"


def fmt_porcentaje(valor, ancho):
    """Formato (4 enteros + punto + 2 decimales) ej: 0000.00, ancho 7."""
    decimales = 2
    enteros = ancho - 1 - decimales
    try:
        v = float(valor) if valor not in (None, "") else 0.0
    except ValueError:
        v = 0.0
    s = f"{v:.{decimales}f}"
    ent, dec = s.split(".")
    ent = ent.zfill(enteros)[-enteros:]
    return f"{ent}.{dec}"


def fmt_fecha(valor, ancho):
    """AAAAMMDD."""
    if valor in (None, ""):
        return " " * ancho
    if isinstance(valor, datetime):
        return valor.strftime("%Y%m%d")
    s = str(valor).strip()
    s = re.sub(r"[^0-9]", "", s)
    return s.ljust(ancho)[:ancho] if s else " " * ancho


def fmt_numerico_simple(valor, ancho):
    if valor in (None, ""):
        return "0".zfill(ancho) if ancho <= 4 else " " * ancho
    s = str(valor).strip()
    s = re.sub(r"[^0-9\-]", "", s)
    if s in ("", "-"):
        s = "0"
    neg = s.startswith("-")
    s = s.lstrip("-")
    s = s.zfill(ancho - 1 if neg else ancho)[: (ancho - 1 if neg else ancho)]
    return ("-" + s) if neg else s


def fmt_alfanumerico(valor, ancho):
    if valor is None:
        valor = ""
    s = str(valor).strip()
    return s.ljust(ancho)[:ancho]


# Campos con formato especial conocido (según notas de la columna OBSERVACIONES)
FORMATOS_ESPECIALES = {
    "F201_CUPO_CREDITO": fmt_decimal_con_signo,
    "F201_PORC_EXCESO_VENTA": fmt_porcentaje,
    "F201_PORC_MIN_MARGEN": fmt_porcentaje,
    "F201_PORC_MAX_MARGEN": fmt_porcentaje,
    "F201_FECHA_INGRESO": fmt_fecha,
    "F201_FECHA_CUPO": fmt_fecha,
    "F201_PORC_TOLERANCIA": fmt_porcentaje,
}


def formatear_campo(campo, valor):
    nombre = campo["nombre"]
    ancho = campo["tamano"]

    if nombre in FORMATOS_ESPECIALES:
        return FORMATOS_ESPECIALES[nombre](valor, ancho)

    if campo["tipo"].startswith("num"):
        return fmt_numerico_simple(valor, ancho)

    return fmt_alfanumerico(valor, ancho)


# ---------------------------------------------------------------------------
# 3. CONSTRUCCIÓN DEL ARCHIVO PLANO
# ---------------------------------------------------------------------------

def leer_datos(path_datos):
    wb = openpyxl.load_workbook(path_datos, data_only=True)
    ws = wb.active
    headers = [normalizar(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    filas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        filas.append(dict(zip(headers, row)))
    return filas


def construir_linea(campos, fila, largo_total, num_consecutivo):
    buffer = [" "] * largo_total
    errores = []

    for campo in campos:
        nombre = campo["nombre"]
        inicio = campo["inicio"]
        ancho = campo["tamano"]

        if nombre == "F_NUMERO_REG":
            valor = num_consecutivo
        elif nombre == "F_TIPO_REG":
            valor = 201
        elif nombre == "F_SUBTIPO_REG":
            valor = 0
        else:
            valor = fila.get(nombre)

        if campo["obligatorio"] == "si" and (valor is None or str(valor).strip() == ""):
            errores.append(f"Campo obligatorio vacío: {campo['nombre_original']}")

        texto = formatear_campo(campo, valor)
        if len(texto) != ancho:
            # seguridad extra: nunca debe desbordar la posición
            texto = texto[:ancho].ljust(ancho)

        pos0 = inicio - 1
        buffer[pos0:pos0 + ancho] = list(texto)

    return "".join(buffer), errores


def generar(path_layout, path_datos, path_salida):
    campos, largo_total = leer_layout(path_layout)
    filas = leer_datos(path_datos)

    lineas = []
    todos_errores = []
    for i, fila in enumerate(filas, start=1):
        linea, errores = construir_linea(campos, fila, largo_total, i)
        lineas.append(linea)
        for e in errores:
            todos_errores.append(f"Fila {i}: {e}")

    with open(path_salida, "w", encoding=ENCODING_SALIDA, newline="\r\n") as f:
        f.write("\n".join(lineas) + "\n")

    print(f"Generadas {len(lineas)} líneas en '{path_salida}' (largo de registro: {largo_total} caracteres).")
    if todos_errores:
        print(f"\n⚠ Se detectaron {len(todos_errores)} advertencias de campos obligatorios vacíos:")
        for e in todos_errores[:50]:
            print("  -", e)
        if len(todos_errores) > 50:
            print(f"  ... y {len(todos_errores) - 50} más.")
    else:
        print("Sin advertencias de campos obligatorios.")


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Uso: python generar_plano_clientes.py layout.xlsx datos.xlsx salida.txt")
        sys.exit(1)
    generar(sys.argv[1], sys.argv[2], sys.argv[3])
