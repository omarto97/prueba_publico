"""
=============================================================
  ANALIZADOR MASIVO DE ARCHIVOS - v2
  Soporta: XLS, XLSX, CSV, TXT, MSG (Outlook)
  Recorre subcarpetas recursivamente
  Genera reporte Excel con scoring de automatizabilidad
=============================================================

INSTALACIÓN:
    pip install "xlrd==1.2.0" openpyxl pandas extract-msg chardet

USO:
    1. Ajusta CARPETA_RAIZ abajo
    2. python analizar_archivos_v2.py

SALIDA:
    reporte_inventario.xlsx  →  5 hojas:
        • Scoring           : puntaje por archivo (ordenado de mayor a menor)
        • Resumen           : una fila por hoja/tabla por archivo
        • Detalle_Nulos     : % nulos por columna
        • Adjuntos_Correos  : adjuntos encontrados en .msg
        • Errores           : archivos que no pudieron leerse
"""

import os
import re
import warnings
import chardet
import pandas as pd
import xlrd
import openpyxl
import datetime as dt

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
#   CONFIGURACIÓN
# ─────────────────────────────────────────────
CARPETA_RAIZ  = "./archivos"        # <-- CAMBIA ESTO
fecha = dt.datetime.now().strftime("%Y%m%d%H%M%S")
name = "reporte_inventario_"+fecha+".xlsx"
ARCHIVO_SALIDA = name

# Extensiones de datos que se analizan dentro de correos y directo
EXTENSIONES_DATOS   = {".xls", ".xlsx", ".csv", ".txt"}
EXTENSIONES_CORREO  = {".msg"}
EXTENSIONES_TODAS   = EXTENSIONES_DATOS | EXTENSIONES_CORREO

# Intentar estos separadores en orden para CSV/TXT con detección automática
SEPARADORES_CANDIDATOS = [",", "\t", ";", "|"]

# Pesos del scoring (deben sumar 100)
PESOS = {
    "estructura_consistente":   20,
    "sin_formulas":             15,
    "pocos_nulos":              20,
    "tipos_consistentes":       15,
    "nombres_columnas_limpios": 10,
    "una_tabla_por_hoja":       10,
    "sin_hojas_vacias":          5,
    "pocas_hojas":               5,
}


# ─────────────────────────────────────────────
#   DESCUBRIMIENTO RECURSIVO DE ARCHIVOS
# ─────────────────────────────────────────────

def descubrir_archivos(carpeta_raiz):
    """
    Recorre recursivamente la carpeta y retorna lista de dicts:
    {ruta, tipo: 'datos'|'correo', origen: 'directo'|'adjunto', correo_padre}
    """
    encontrados = []

    for dirpath, _, filenames in os.walk(carpeta_raiz):
        for filename in filenames:
            if filename.startswith("~$"):   # archivos temporales de Office
                continue
            ext = os.path.splitext(filename)[1].lower()
            ruta = os.path.join(dirpath, filename)

            if ext in EXTENSIONES_DATOS:
                encontrados.append({
                    "ruta": ruta,
                    "tipo": "datos",
                    "origen": "directo",
                    "correo_padre": None,
                })
            elif ext in EXTENSIONES_CORREO:
                encontrados.append({
                    "ruta": ruta,
                    "tipo": "correo",
                    "origen": "directo",
                    "correo_padre": None,
                })

    return encontrados


# ─────────────────────────────────────────────
#   MANEJO DE CORREOS .MSG
# ─────────────────────────────────────────────

def extraer_adjuntos_msg(ruta_msg):
    """
    Extrae adjuntos de un .msg de Outlook.
    Retorna lista de dicts con info de cada adjunto.
    """
    try:
        import extract_msg
    except ImportError:
        return [], "Instala extract-msg: pip install extract-msg"

    adjuntos_info = []
    datos_extraidos = []   # (nombre, bytes) de adjuntos que son datos

    try:
        msg = extract_msg.Message(ruta_msg)
        asunto  = msg.subject or "(sin asunto)"
        remitente = msg.sender or "?"
        fecha   = str(msg.date or "?")

        for att in msg.attachments:
            nombre_adj = att.longFilename or att.shortFilename or "adjunto_sin_nombre"
            ext_adj    = os.path.splitext(nombre_adj)[1].lower()
            datos_bytes = att.data if hasattr(att, "data") else None
            tamanio_kb  = round(len(datos_bytes) / 1024, 1) if datos_bytes else None

            es_dato = ext_adj in EXTENSIONES_DATOS

            adjuntos_info.append({
                "correo":       os.path.basename(ruta_msg),
                "ruta_correo":  ruta_msg,
                "asunto":       asunto,
                "remitente":    remitente,
                "fecha_correo": fecha,
                "adjunto":      nombre_adj,
                "extension":    ext_adj,
                "tamanio_kb":   tamanio_kb,
                "es_dato":      es_dato,
            })

            if es_dato and datos_bytes:
                datos_extraidos.append((nombre_adj, ext_adj, datos_bytes))

        msg.close()
    except Exception as e:
        return [], str(e)

    return adjuntos_info, datos_extraidos


# ─────────────────────────────────────────────
#   DETECCIÓN DE ENCODING Y SEPARADOR CSV/TXT
# ─────────────────────────────────────────────

def detectar_encoding(datos_bytes=None, ruta=None):
    """Detecta encoding usando chardet."""
    try:
        muestra = datos_bytes[:50_000] if datos_bytes else open(ruta, "rb").read(50_000)
        resultado = chardet.detect(muestra)
        enc = resultado.get("encoding") or "utf-8"
        # Normalizar variantes comunes
        enc = enc.replace("ISO-8859-1", "latin-1").replace("Windows-1252", "cp1252")
        return enc
    except Exception:
        return "utf-8"


def detectar_separador(datos_texto, n_lineas=5):
    """
    Detecta el separador más probable probando los candidatos en las
    primeras n_lineas. Gana el que produce más columnas consistentes.
    """
    lineas = [l for l in datos_texto.splitlines()[:n_lineas] if l.strip()]
    if not lineas:
        return ","

    mejor_sep  = ","
    mejor_cols = 0

    for sep in SEPARADORES_CANDIDATOS:
        conteos = [len(l.split(sep)) for l in lineas]
        if len(set(conteos)) == 1 and conteos[0] > mejor_cols:
            mejor_cols = conteos[0]
            mejor_sep  = sep

    return mejor_sep


def leer_csv_txt(datos_bytes=None, ruta=None, nombre=""):
    """
    Lee un CSV o TXT con detección automática de encoding y separador.
    Retorna DataFrame o lanza excepción.
    """
    encoding = detectar_encoding(datos_bytes=datos_bytes, ruta=ruta)

    if datos_bytes:
        texto = datos_bytes.decode(encoding, errors="replace")
    else:
        with open(ruta, "r", encoding=encoding, errors="replace") as f:
            texto = f.read()

    sep = detectar_separador(texto)

    import io
    df = pd.read_csv(io.StringIO(texto), sep=sep, dtype=str, on_bad_lines="skip")
    return df, sep, encoding


# ─────────────────────────────────────────────
#   UTILIDADES DE ANÁLISIS
# ─────────────────────────────────────────────

def es_nombre_columna_limpio(nombre):
    nombre = str(nombre).strip()
    if nombre.startswith("Unnamed") or nombre == "":
        return False
    if re.search(r"[^a-zA-Z0-9áéíóúÁÉÍÓÚñÑüÜ _\-./()]", nombre):
        return False
    return True


def detectar_formulas_xls(ruta=None, wb=None):
    count = 0
    try:
        libro = wb or xlrd.open_workbook(ruta, formatting_info=False)
        for sheet in libro.sheets():
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    cell = sheet.cell(row, col)
                    if cell.ctype == xlrd.XL_CELL_TEXT and str(cell.value).startswith("="):
                        count += 1
    except Exception:
        pass
    return count


def detectar_formulas_xlsx(ruta=None, wb=None):
    count = 0
    try:
        libro = wb or openpyxl.load_workbook(ruta, read_only=True, data_only=False)
        for ws in libro.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == "f" or (
                        isinstance(cell.value, str) and cell.value.startswith("=")
                    ):
                        count += 1
        if not wb:
            libro.close()
    except Exception:
        pass
    return count


def detectar_celdas_fusionadas(ruta, es_xlsx):
    count = 0
    try:
        if es_xlsx:
            wb = openpyxl.load_workbook(ruta, read_only=False, data_only=True)
            for ws in wb.worksheets:
                count += len(ws.merged_cells.ranges)
            wb.close()
        else:
            wb = xlrd.open_workbook(ruta, formatting_info=True)
            for sheet in wb.sheets():
                count += len(sheet.merged_cells())
    except Exception:
        pass
    return count


def primera_fila_con_datos(df_raw):
    for i, row in df_raw.iterrows():
        if row.notna().any():
            return i
    return 0


def analizar_df(df, nombre_hoja="Hoja1"):
    """Analiza un DataFrame ya cargado y retorna dict con métricas."""
    hoja = {
        "nombre_hoja":          nombre_hoja,
        "vacia":                True,
        "fila_encabezado":      0,
        "encabezado_en_fila1":  True,
        "datos_desde_A1":       True,
        "num_columnas":         0,
        "nombres_columnas":     [],
        "columnas_limpias":     0,
        "pct_columnas_limpias": 0.0,
        "filas_totales":        0,
        "filas_con_datos":      0,
        "filas_en_blanco":      0,
        "pct_nulos_por_columna": {},
        "tipos_por_columna":    {},
        "columnas_tipo_mixto":  [],
    }

    if df is None or df.empty:
        return hoja

    df_sin = df.dropna(how="all")
    if df_sin.empty:
        return hoja

    hoja["vacia"]           = False
    hoja["filas_totales"]   = df.shape[0]
    hoja["filas_con_datos"] = df_sin.shape[0]
    hoja["filas_en_blanco"] = df.shape[0] - df_sin.shape[0]
    hoja["num_columnas"]    = len(df.columns)
    hoja["nombres_columnas"] = [str(c) for c in df.columns]

    limpias = [c for c in df.columns if es_nombre_columna_limpio(c)]
    hoja["columnas_limpias"]      = len(limpias)
    hoja["pct_columnas_limpias"]  = round(len(limpias) / max(len(df.columns), 1) * 100, 1)
    hoja["pct_nulos_por_columna"] = df.isnull().mean().mul(100).round(1).to_dict()

    tipos  = {}
    mixtas = []
    for col in df.columns:
        serie = df[col].dropna()
        if serie.empty:
            tipos[str(col)] = "vacío"
            continue
        tipo_pd = str(df[col].dtype)
        if "int" in tipo_pd or "float" in tipo_pd:
            tipos[str(col)] = "numérico"
        elif "datetime" in tipo_pd:
            tipos[str(col)] = "fecha"
        elif "bool" in tipo_pd:
            tipos[str(col)] = "booleano"
        else:
            tiene_nums  = pd.to_numeric(serie, errors="coerce").notna().sum()
            tiene_texto = serie.apply(lambda x: isinstance(x, str)).sum()
            if tiene_nums > 0 and tiene_texto > 0:
                tipos[str(col)] = "mixto"
                mixtas.append(str(col))
            else:
                tipos[str(col)] = "texto"

    hoja["tipos_por_columna"]   = tipos
    hoja["columnas_tipo_mixto"] = mixtas
    return hoja


# ─────────────────────────────────────────────
#   ANÁLISIS POR TIPO DE ARCHIVO
# ─────────────────────────────────────────────

def analizar_excel(ruta=None, datos_bytes=None, nombre="archivo.xlsx"):
    es_xlsx = nombre.lower().endswith(".xlsx")
    engine  = "openpyxl" if es_xlsx else "xlrd"

    resultado = {
        "archivo":                nombre,
        "ruta":                   ruta or "(adjunto en correo)",
        "formato":                "xlsx" if es_xlsx else "xls",
        "tamanio_kb":             round(len(datos_bytes) / 1024, 1) if datos_bytes else (
                                  round(os.path.getsize(ruta) / 1024, 1) if ruta else None),
        "num_hojas":              0,
        "total_formulas":         0,
        "total_celdas_fusionadas": 0,
        "separador":              "N/A",
        "encoding":               "N/A",
        "hojas":                  [],
        "error":                  None,
    }

    try:
        import io
        src = io.BytesIO(datos_bytes) if datos_bytes else ruta

        resultado["total_formulas"]          = detectar_formulas_xlsx(src) if es_xlsx else detectar_formulas_xls(src)
        resultado["total_celdas_fusionadas"] = detectar_celdas_fusionadas(src, es_xlsx) if ruta else 0

        hojas_raw = pd.read_excel(src, sheet_name=None, header=None, engine=engine)
        resultado["num_hojas"] = len(hojas_raw)

        for nombre_hoja, df_raw in hojas_raw.items():
            header_row = primera_fila_con_datos(df_raw)
            src2 = io.BytesIO(datos_bytes) if datos_bytes else ruta
            df   = pd.read_excel(src2, sheet_name=nombre_hoja, header=header_row, engine=engine)
            info = analizar_df(df, nombre_hoja)
            info["fila_encabezado"]     = int(header_row)
            info["encabezado_en_fila1"] = header_row == 0
            info["datos_desde_A1"]      = header_row == 0
            resultado["hojas"].append(info)

    except Exception as e:
        resultado["error"] = str(e)

    return resultado


def analizar_csv_txt(ruta=None, datos_bytes=None, nombre="archivo.csv"):
    resultado = {
        "archivo":                nombre,
        "ruta":                   ruta or "(adjunto en correo)",
        "formato":                os.path.splitext(nombre)[1].lower().lstrip("."),
        "tamanio_kb":             round(len(datos_bytes) / 1024, 1) if datos_bytes else (
                                  round(os.path.getsize(ruta) / 1024, 1) if ruta else None),
        "num_hojas":              1,
        "total_formulas":         0,
        "total_celdas_fusionadas": 0,
        "separador":              "?",
        "encoding":               "?",
        "hojas":                  [],
        "error":                  None,
    }

    try:
        df, sep, enc = leer_csv_txt(datos_bytes=datos_bytes, ruta=ruta, nombre=nombre)
        resultado["separador"] = repr(sep)
        resultado["encoding"]  = enc
        info = analizar_df(df, "datos")
        info["encabezado_en_fila1"] = True
        info["datos_desde_A1"]      = True
        info["fila_encabezado"]     = 0
        resultado["hojas"].append(info)
    except Exception as e:
        resultado["error"] = str(e)

    return resultado


def analizar_archivo_datos(ruta=None, datos_bytes=None, nombre=None):
    """Dispatcher según extensión."""
    nombre = nombre or os.path.basename(ruta or "")
    ext = os.path.splitext(nombre)[1].lower()
    if ext in (".xls", ".xlsx"):
        return analizar_excel(ruta=ruta, datos_bytes=datos_bytes, nombre=nombre)
    elif ext in (".csv", ".txt"):
        return analizar_csv_txt(ruta=ruta, datos_bytes=datos_bytes, nombre=nombre)
    else:
        return {"archivo": nombre, "error": f"Extensión no soportada: {ext}", "hojas": []}


# ─────────────────────────────────────────────
#   SCORING
# ─────────────────────────────────────────────

def calcular_score(analisis):
    score = {}
    hojas = analisis.get("hojas", [])
    hojas_con_datos = [h for h in hojas if not h.get("vacia", True)]

    if not hojas_con_datos:
        return {k: 0 for k in PESOS}, 0, "🔴 Muy baja"

    pct_enc_ok       = sum(1 for h in hojas_con_datos if h.get("encabezado_en_fila1")) / len(hojas_con_datos)
    pen_fusion        = min(1.0, analisis.get("total_celdas_fusionadas", 0) / 20)
    score["estructura_consistente"] = round(
        PESOS["estructura_consistente"] * pct_enc_ok * (1 - pen_fusion * 0.5)
    )

    pen_form = min(1.0, analisis.get("total_formulas", 0) / 50)
    score["sin_formulas"] = round(PESOS["sin_formulas"] * (1 - pen_form))

    todos_nulos = []
    for h in hojas_con_datos:
        todos_nulos.extend(h.get("pct_nulos_por_columna", {}).values())
    prom_nulos = sum(todos_nulos) / max(len(todos_nulos), 1)
    score["pocos_nulos"] = round(PESOS["pocos_nulos"] * max(0, 1 - prom_nulos / 100))

    total_cols  = sum(h.get("num_columnas", 0) for h in hojas_con_datos)
    cols_mixtas = sum(len(h.get("columnas_tipo_mixto", [])) for h in hojas_con_datos)
    score["tipos_consistentes"] = round(PESOS["tipos_consistentes"] * (1 - cols_mixtas / max(total_cols, 1)))

    prom_limpias = sum(h.get("pct_columnas_limpias", 0) for h in hojas_con_datos) / len(hojas_con_datos)
    score["nombres_columnas_limpios"] = round(PESOS["nombres_columnas_limpios"] * prom_limpias / 100)

    pct_A1 = sum(1 for h in hojas_con_datos if h.get("datos_desde_A1")) / len(hojas_con_datos)
    score["una_tabla_por_hoja"] = round(PESOS["una_tabla_por_hoja"] * pct_A1)

    hojas_vacias = sum(1 for h in hojas if h.get("vacia", True))
    score["sin_hojas_vacias"] = round(PESOS["sin_hojas_vacias"] * (1 - min(1.0, hojas_vacias / max(len(hojas), 1))))

    puntos_hojas = max(0, 1 - (analisis.get("num_hojas", 1) - 1) / 10)
    score["pocas_hojas"] = round(PESOS["pocas_hojas"] * puntos_hojas)

    total = sum(score.values())
    if   total >= 80: cat = "🟢 Alta"
    elif total >= 55: cat = "🟡 Media"
    elif total >= 30: cat = "🟠 Baja"
    else:             cat = "🔴 Muy baja"

    return score, total, cat


# ─────────────────────────────────────────────
#   EXPORTAR REPORTE
# ─────────────────────────────────────────────

def exportar_reporte(resultados, adjuntos_correo, ruta_salida):
    filas_resumen = []
    filas_scoring = []
    filas_nulos   = []
    filas_errores = []

    for r in resultados:
        if r.get("error") and not r.get("hojas"):
            filas_errores.append({
                "archivo": r.get("archivo"), "ruta": r.get("ruta"),
                "formato": r.get("formato", "?"), "error": r["error"],
            })
            continue

        score_d, score_t, cat = calcular_score(r)

        filas_scoring.append({
            "archivo":                  r["archivo"],
            "ruta":                     r.get("ruta", ""),
            "formato":                  r.get("formato", "?"),
            "origen":                   r.get("origen", "directo"),
            "correo_padre":             r.get("correo_padre", ""),
            "score_total (0-100)":      score_t,
            "categoria":                cat,
            **{k: score_d.get(k) for k in PESOS},
            "num_hojas":                r.get("num_hojas", 1),
            "total_formulas":           r.get("total_formulas", 0),
            "total_celdas_fusionadas":  r.get("total_celdas_fusionadas", 0),
            "separador":                r.get("separador", "N/A"),
            "encoding":                 r.get("encoding", "N/A"),
            "tamanio_kb":               r.get("tamanio_kb"),
        })

        for h in r.get("hojas", []):
            filas_resumen.append({
                "archivo":              r["archivo"],
                "origen":               r.get("origen", "directo"),
                "correo_padre":         r.get("correo_padre", ""),
                "formato":              r.get("formato", "?"),
                "tamanio_kb":           r.get("tamanio_kb"),
                "hoja":                 h["nombre_hoja"],
                "num_hojas_total":      r.get("num_hojas", 1),
                "vacia":                h.get("vacia"),
                "fila_encabezado":      h.get("fila_encabezado"),
                "encabezado_en_fila1":  h.get("encabezado_en_fila1"),
                "datos_desde_A1":       h.get("datos_desde_A1"),
                "num_columnas":         h.get("num_columnas"),
                "filas_con_datos":      h.get("filas_con_datos"),
                "filas_en_blanco":      h.get("filas_en_blanco"),
                "pct_columnas_limpias": h.get("pct_columnas_limpias"),
                "columnas_tipo_mixto":  " | ".join(h.get("columnas_tipo_mixto", [])),
                "formulas_libro":       r.get("total_formulas", 0),
                "celdas_fusionadas":    r.get("total_celdas_fusionadas", 0),
                "columnas":             " | ".join(h.get("nombres_columnas", [])),
            })

            for col, pct in h.get("pct_nulos_por_columna", {}).items():
                filas_nulos.append({
                    "archivo":   r["archivo"],
                    "hoja":      h["nombre_hoja"],
                    "columna":   col,
                    "tipo_dato": h.get("tipos_por_columna", {}).get(col, "?"),
                    "pct_nulos": pct,
                })

    df_scoring  = pd.DataFrame(filas_scoring).sort_values("score_total (0-100)", ascending=False)
    df_resumen  = pd.DataFrame(filas_resumen)
    df_nulos    = pd.DataFrame(filas_nulos)
    df_adjuntos = pd.DataFrame(adjuntos_correo)
    df_errores  = pd.DataFrame(filas_errores)

    COLORES_CAT = {
        "🟢 Alta":     "C6EFCE",
        "🟡 Media":    "FFEB9C",
        "🟠 Baja":     "FFCC99",
        "🔴 Muy baja": "FFC7CE",
    }

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        df_scoring.to_excel(writer,  sheet_name="Scoring",          index=False)
        df_resumen.to_excel(writer,  sheet_name="Resumen",           index=False)
        df_nulos.to_excel(writer,    sheet_name="Detalle_Nulos",     index=False)
        df_adjuntos.to_excel(writer, sheet_name="Adjuntos_Correos",  index=False)
        df_errores.to_excel(writer,  sheet_name="Errores",           index=False)

        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = writer.book

        # ── Formato hoja Scoring ──
        ws = wb["Scoring"]
        for col_cells in ws.columns:
            ancho = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(ancho + 4, 45)

        cat_col_idx = None
        for cell in ws[1]:
            if cell.value == "categoria":
                cat_col_idx = cell.column
                break

        if cat_col_idx:
            for row in ws.iter_rows(min_row=2):
                val   = str(row[cat_col_idx - 1].value or "")
                color = COLORES_CAT.get(val, "FFFFFF")
                fill  = PatternFill("solid", fgColor=color)
                for cell in row:
                    cell.fill = fill

        # ── Header en negrita en todas las hojas ──
        for ws_name in ["Scoring", "Resumen", "Detalle_Nulos", "Adjuntos_Correos", "Errores"]:
            if ws_name in wb.sheetnames:
                for cell in wb[ws_name][1]:
                    cell.font      = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")

    print(f"\n✅  Reporte guardado: {ruta_salida}")


# ─────────────────────────────────────────────
#   MAIN
# ─────────────────────────────────────────────

def main():
    if not os.path.isdir(CARPETA_RAIZ):
        print(f"❌  Carpeta no encontrada: {CARPETA_RAIZ}")
        return

    print(f"🔍  Escaneando: {CARPETA_RAIZ}\n")
    encontrados = descubrir_archivos(CARPETA_RAIZ)

    directos = [f for f in encontrados if f["tipo"] == "datos"]
    correos  = [f for f in encontrados if f["tipo"] == "correo"]

    print(f"   Archivos de datos  : {len(directos)}")
    print(f"   Correos .msg       : {len(correos)}\n")

    resultados        = []
    adjuntos_correo   = []
    contador          = 0
    total_estimado    = len(directos) + len(correos)

    # ── Archivos de datos directos ──
    for item in directos:
        contador += 1
        nombre = os.path.basename(item["ruta"])
        print(f"  [{contador:>4}/{total_estimado}] {nombre[:60]:<60}", end=" ", flush=True)
        r = analizar_archivo_datos(ruta=item["ruta"])
        r["origen"]       = "directo"
        r["correo_padre"] = None
        estado = f"❌ {r.get('error','')[:40]}" if r.get("error") and not r.get("hojas") else f"✔  {r.get('num_hojas',1)} hojas"
        print(estado)
        resultados.append(r)

    # ── Correos .msg ──
    for item in correos:
        contador += 1
        nombre_correo = os.path.basename(item["ruta"])
        print(f"  [{contador:>4}/{total_estimado}] 📧 {nombre_correo[:57]:<57}", end=" ", flush=True)

        info_adjuntos, datos_extraidos = extraer_adjuntos_msg(item["ruta"])

        if isinstance(datos_extraidos, str):
            # Es un mensaje de error
            print(f"❌ {datos_extraidos[:50]}")
            resultados.append({"archivo": nombre_correo, "ruta": item["ruta"],
                                "formato": "msg", "error": datos_extraidos, "hojas": []})
            continue

        adjuntos_correo.extend(info_adjuntos)
        n_datos = len([a for a in info_adjuntos if a["es_dato"]])
        print(f"✔  {len(info_adjuntos)} adjuntos ({n_datos} datos)")

        for nombre_adj, ext_adj, bytes_adj in datos_extraidos:
            print(f"         ↳ analizando adjunto: {nombre_adj[:50]}", end=" ", flush=True)
            r = analizar_archivo_datos(datos_bytes=bytes_adj, nombre=nombre_adj)
            r["origen"]       = "adjunto"
            r["correo_padre"] = nombre_correo
            estado = f"❌ {r.get('error','')[:35]}" if r.get("error") and not r.get("hojas") else f"✔  {r.get('num_hojas',1)} hojas"
            print(estado)
            resultados.append(r)

    print(f"\n📊  Generando reporte ...\n")
    exportar_reporte(resultados, adjuntos_correo, ARCHIVO_SALIDA)

    # Resumen final en consola
    con_score   = [r for r in resultados if not (r.get("error") and not r.get("hojas"))]
    scores      = [calcular_score(r)[1] for r in con_score]
    categorias  = [calcular_score(r)[2] for r in con_score]

    print(f"\n   📁  Archivos procesados   : {len(resultados)}")
    print(f"   ✅  Analizados con éxito  : {len(con_score)}")
    print(f"   ❌  Con errores           : {len(resultados) - len(con_score)}")
    if scores:
        print(f"   📈  Score promedio        : {round(sum(scores)/len(scores), 1)}")
        print(f"   🟢  Alta automatizabilidad: {categorias.count('🟢 Alta')}")
        print(f"   🟡  Media                 : {categorias.count('🟡 Media')}")
        print(f"   🟠  Baja                  : {categorias.count('🟠 Baja')}")
        print(f"   🔴  Muy baja              : {categorias.count('🔴 Muy baja')}")


if __name__ == "__main__":
    main()
